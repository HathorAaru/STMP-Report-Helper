from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .name_matching import display_name, name_key


@dataclass
class StudentRecord:
    name: str
    attendance_actual: float | None = None
    attendance_authorised: float | None = None
    attendance_unauthorised: float | None = None
    reading: str | None = None
    writing: str | None = None
    maths: str | None = None


def _clean_header(value):
    return " ".join(str(value or "").split())


def _find_header_indexes(row):
    return {_clean_header(value): index for index, value in enumerate(row) if value is not None}


def load_attendance(path):
    workbook = load_workbook(Path(path), data_only=True)
    sheet = workbook["Report Data"] if "Report Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = _find_header_indexes(rows[0])

    required = [
        "Student",
        "Present R/C: Marks (%)",
        "Auth. Absent R/C: Marks (%)",
        "Unauth. Absent R/C: Marks (%)",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Attendance spreadsheet is missing columns: {', '.join(missing)}")

    records = {}
    for row in rows[1:]:
        raw_name = row[headers["Student"]]
        if not raw_name:
            continue
        records[name_key(raw_name)] = {
            "name": display_name(raw_name),
            "attendance_actual": row[headers["Present R/C: Marks (%)"]],
            "attendance_authorised": row[headers["Auth. Absent R/C: Marks (%)"]],
            "attendance_unauthorised": row[headers["Unauth. Absent R/C: Marks (%)"]],
        }
    return records


def _find_tracker_header_row(rows):
    for row_index, row in enumerate(rows):
        headers = _find_header_indexes(row)
        if {"Name", "Reading", "Writing", "Overall Maths judgement"}.issubset(headers):
            return row_index, row_index + 1, headers

    assessment_headers = None
    for row_index, row in enumerate(rows):
        headers = _find_header_indexes(row)
        if {"Reading", "Writing", "Overall Maths judgement"}.issubset(headers):
            assessment_headers = headers
            break

    if assessment_headers:
        for row_index, row in enumerate(rows):
            headers = _find_header_indexes(row)
            if "Name" in headers:
                combined_headers = dict(assessment_headers)
                combined_headers["Name"] = headers["Name"]
                return row_index, row_index + 1, combined_headers

    raise ValueError("Could not find the tracker header row with Name, Reading, Writing, and Overall Maths judgement.")


def load_attainment(path):
    workbook = load_workbook(Path(path), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    _, first_data_row_index, headers = _find_tracker_header_row(rows)

    records = {}
    for row in rows[first_data_row_index:]:
        raw_name = row[headers["Name"]]
        if not raw_name:
            continue
        records[name_key(raw_name)] = {
            "name": display_name(raw_name),
            "reading": row[headers["Reading"]],
            "writing": row[headers["Writing"]],
            "maths": row[headers["Overall Maths judgement"]],
        }
    return records


def load_students(attendance_path, attainment_path):
    attendance = load_attendance(attendance_path)
    attainment = load_attainment(attainment_path)
    all_keys = sorted(set(attendance) | set(attainment), key=lambda key: (attendance.get(key) or attainment.get(key))["name"])

    students = []
    warnings = []
    for key in all_keys:
        combined = {}
        if key in attendance:
            combined.update(attendance[key])
        else:
            combined.update({"name": attainment[key]["name"]})
            warnings.append(f"Attendance missing for {combined['name']}")

        if key in attainment:
            combined.update(attainment[key])
        else:
            warnings.append(f"Attainment missing for {combined['name']}")

        students.append(StudentRecord(**combined))
    return students, warnings
